import os
import re
import shutil

import torch
from pathlib import Path
from openpyxl import Workbook, load_workbook


def load_infer_model(generator, path, device, key: str | None):
    if not path.exists():
        raise FileNotFoundError(f"Model does not exist.")

    ckpt = torch.load(path, map_location=device)
    state_dict = ckpt if key is None else ckpt[key]
    generator.load_state_dict(state_dict)


def load_srgan_checkpoint(generator, discriminator, optimizer_G=None, optimizer_D=None, device="cuda"):
    project_root = Path(__file__).resolve().parents[2]
    srgan_folder_path = project_root / 'outputs' / 'srgan'

    if not os.path.exists(srgan_folder_path):
        return 0

    checkpoints = list(srgan_folder_path.glob("srgan_step*.pth"))

    def extract_step(path):
        m = re.search(r"srgan_step(\d+)\.pth", path.name)
        return int(m.group(1)) if m else -1

    checkpoint_path = max(checkpoints, key=extract_step)
    checkpoint = torch.load(checkpoint_path, map_location=device)

    generator.load_state_dict(checkpoint["generator_state_dict"])
    discriminator.load_state_dict(checkpoint["discriminator_state_dict"])

    # Load optimizers
    if optimizer_G is not None and "optimizer_G" in checkpoint:
        optimizer_G.load_state_dict(checkpoint["optimizer_G"])
    if optimizer_D is not None and "optimizer_D" in checkpoint:
        optimizer_D.load_state_dict(checkpoint["optimizer_D"])

    step = int(checkpoint.get("step", 0))

    return step


def load_ptg_checkpoint(generator, optimizer, opt):
    project_root = Path(__file__).resolve().parents[2]
    ptg_folder_path = project_root / 'outputs' / 'generator_pretrained'

    if not os.path.exists(ptg_folder_path):
        return 0

    checkpoints = list(ptg_folder_path.glob("ptg_*.pth"))

    def extract_step(path):
        m = re.search(r"ptg_(\d+)\.pth", path.name)
        return int(m.group(1)) if m else -1

    checkpoint_path = max(checkpoints, key=extract_step)
    print(checkpoint_path)
    checkpoint = torch.load(checkpoint_path, map_location=opt.device)

    generator.load_state_dict(checkpoint["generator_state_dict"])

    # Load optimizer
    if optimizer is not None and "optimizer" in checkpoint:
        optimizer.load_state_dict(checkpoint["optimizer"])

    step = int(checkpoint.get("step", 0))

    return step


def init_log():
    log_path = Path("outputs/logs/srgan_metrics.xlsx")
    log_path.parent.mkdir(parents=True, exist_ok=True)

    if log_path.exists():
        wb = load_workbook(log_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append([
            "step",
            "loss_D",
            "content_loss",
            "adversarial_loss",
            "loss_G"
        ])
        wb.save(log_path)

    return wb, ws, log_path


def delete_experiment_artifacts():
    project_root = Path(__file__).resolve().parents[2]

    path_outputs = project_root / 'outputs'
    path_models = project_root / 'models'

    if os.path.exists(path_outputs):
        shutil.rmtree(path_outputs)

    if os.path.exists(path_models):
        shutil.rmtree(path_models)
